import openpyxl
from datetime import date

import ProcessDBRecords
import LookupFunctions

import clsLogEntry
import UADW_GRAD_AVNS_ADDR0
import UniversityToBiologyTranslation
import DeleteFunctions

DryRun = False

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


def readfile():
    # ## getting the reference of the cells which we want to get the data from
    # name = sheet['A1']
    # tag = sheet.cell(row = 1, column = 2)
    #
    # ## printing the values of cells
    # print(name.value)
    # print(tag.value)

    Process_UADW_GRAD_AVNS_ADDR_Spreadsheet("C:\\Temp\\UADW_GRAD_AVNS_ADDR_1229412348.xlsx", DryRun)


def Process_UADW_GRAD_AVNS_ADDR_Spreadsheet(SpreadsheetFilepath, DryRun):
    #
    #  UADW_GRAD_AVNS_ADDR provides information on info on graduate students
    #
    #    Degree area and type
    #    Advisors
    #    Date of Admission
    #    Mailing Address
    #    Telephone
    #
    #  Module uses info from Excel spreadsheet to update or reconcile student records.  Update will add new grad student information if grad
    #  student not found in database; it will overwrite existing information in database about that grad student with the University's information.
    #
    #  Reconciliation is similar but also logs changes to existing grad student records.  It is under development.
    #
    # opening the previously created xlsx file using load_workbook()' method

    xlsx = openpyxl.load_workbook(SpreadsheetFilepath)

    # getting the sheet to active
    sheet = xlsx.active

    # This is the standard format in University PeopleSoft spreadsheets
    TitleRow = 1
    HeaderRow = 2
    DataStartingRow = 3

    # Number of rows and columns
    MaxRows = sheet.max_row
    MaxColumns = sheet.max_column

    # Get sheet title
    sheettitle = sheet['A1']

    # Load data into list of dict objects
    List1 = GetDictionaryListFromExcelFile(sheet, HeaderRow, DataStartingRow, MaxRows, MaxColumns)

    for datadict in List1:
        Process_UADW_GRAD_AVNS_ADDR_RowDict(SpreadsheetFilepath, datadict, DryRun)




def Process_UADW_GRAD_AVNS_ADDR_RowDict(SpreadsheetFilepath, DictItem, DryRun):

    # Fixed constant values
    DatasourceID = 1  # "1" == "University Data Source"
    ConfidentialityID = 2  # "2" == "Sensitive"
    EmailTypeID = 2  # "2" == "University Email"

    # If graduate student is not recognized by EmpID, create new grad student account

    if not LookupFunctions.IsEmpIDInBiologyDatabase(DictItem["ID"]):

        # Only create entry if DryRun == False
        if not DryRun:

            # Add new person record for graduate student
            UADW_GRAD_AVNS_ADDR0.addgradstudentrecordfrom_UADW_GRAD_AVNS_ADDR_RowDict(DictItem)

            # Get newly created graduate student personid
            gradstudentid = LookupFunctions.GetPersonIDByLastnameFirstname(DictItem["Last"], DictItem["First Name"])

        # Log creation of new grad student
        UADW_GRAD_AVNS_ADDR0.addlogentry(SpreadsheetFilepath, DictItem["First Name"], DictItem["Last"], \
                                         "Added new Graduate Student person", \
                                         str(gradstudentid))

    # If graduate student is recognized by EmpID, update grad student account
    else:

        # Get created graduate student personid
        gradstudentid = LookupFunctions.GetPersonIDByLastnameFirstname(DictItem["Last"], DictItem["First Name"])

        # Log creation of new grad student
        UADW_GRAD_AVNS_ADDR0.addlogentry(SpreadsheetFilepath, \
                                         DictItem["First Name"], \
                                         DictItem["Last"], \
                                         str(gradstudentid), \
                                         "Updating existing Graduate Student person", \
                                         "")

        # Update grad student information
        if not DryRun:

            # 01 - Set Biology Degree Area / Type
            if not DictItem["Plan10"] == "":

                # Return tuple of (degree area, degree type id)
                DegreeAreaType = UniversityToBiologyTranslation.GetBIODBDegreeAreaAndType(DictItem["Plan10"])

                if DegreeAreaType[0] == -1: # Error
                    UADW_GRAD_AVNS_ADDR0.addlogentry(SpreadsheetFilepath, \
                                         DictItem["First Name"], \
                                         DictItem["Last"], \
                                         str(gradstudentid), \
                                         "*ERROR* Cannot determine degree area '{0}' of graduate student \
                                         '{1} {2}'".format(DictItem["Plan10"], DictItem["First Name"], DictItem["Last"]), \
                                         "")
                else:
                    StartingSemesterID = UniversityToBiologyTranslation.GetBIODBSemesterID(DictItem["Admit Term"])

                    if not DryRun:
                        DeleteFunctions.DeletePersonDegreeAreasAndTypes(gradstudentid, DatasourceID)
                        AddPersonDegreeAreaAndType(gradstudentid, DegreeAreaType[0], DegreeAreaType[1], StartingSemesterID, DatasourceID)

                    note1 = "Updated biology DegreeArea to '{0}', DegreeType to '{1}', StartingSemesterID to '{2}'".format(DegreeAreaType[0], DegreeAreaType[1], StartingSemesterID)
                    UADW_GRAD_AVNS_ADDR0.addlogentry(SpreadsheetFilepath, \
                                                     DictItem["First Name"], \
                                                     DictItem["Last"], \
                                                     str(gradstudentid), \
                                                     note1, \
                                                     "")
            # 02 - Set Advisor
            # 03 - Set Address
            # 04 - Set Email
            # 05 - Set Home Phones




        else:
            print("Update new student account")


# def GetDictionaryListFromNewAccessDBPeopleTable(Database):
#
#     mdb = 'c:/temp/BIO DEPT.mdb'
#     drv = '{Microsoft Access Driver (*.mdb)}'
#     pwd = 'pw'
#
#     # connect to access db
#     # con = pyodbc.connect('DRIVER={};DBQ={};PWD={}'.format(drv, mdb, pwd))
#     con = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=c:\Temp\BIO DEPT.mdb;')
#     cur = con.cursor()
#
#     sql = "SELECT * FROM tblgsPROFILE;"
#
#     try:
#         rows = cur.execute(sql).fetchall()
#     except Exception as e:
#         print(e)
#
#     cur.close()
#     con.close()
#
#     return rows


def GetDictionaryListFromExcelFile(Spreadsheet, HeaderRow, StartingDataRow, MaxDataRows, MaxColumns):
    ListOfData = []

    for rowpointer in range(StartingDataRow, MaxDataRows):
        ListOfData.append(ConvertSpreadsheetRowToDict(Spreadsheet, HeaderRow, rowpointer, MaxColumns))

    return ListOfData


def ConvertSpreadsheetRowToDict(Worksheet, HeaderRow, DataRow, DataColumns):
    SpreadsheetDict = {}

    for colpointer in range(1, DataColumns):
        keytext = Worksheet.cell(row=HeaderRow, column=colpointer).value
        valuetext = Worksheet.cell(row=DataRow, column=colpointer).value

        SpreadsheetDict[keytext] = valuetext

    return SpreadsheetDict


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')

    # fields = ["id", "countryname"]
    #
    #
    #
    # p1 = ProcessDBRecords.ExecuteSQLCommand("select * from data_persons where ualbanyempid = '001290047'", fields)
    # rows = p1.ExecuteSQLCommandReturnRows()
    # p1.printsqlstring()
    # p1.printfields()
    # del p1
    #
    # p1 = ProcessDBRecords.ExecuteSQLCommand("select * from code_countries", fields)
    # list1 = p1.ExecuteSQLCommandReturnDictionaryCollection()
    #
    # p1.printfields()
    # p1.printfields()







    readfile()
